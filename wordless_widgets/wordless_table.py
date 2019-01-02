#
# Wordless: Table
#
# Copyright (C) 2018-2019 Ye Lei (叶磊) <blkserene@gmail.com>
#
# License Information: https://github.com/BLKSerene/Wordless/blob/master/LICENSE.txt
#

import csv
import os

from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *

import numpy
import openpyxl

from wordless_widgets import wordless_box, wordless_message_box
from wordless_utils import wordless_conversion

class Wordless_Label_Html(QLabel):
    def __init__(self, label, parent):
        super().__init__(label, parent)

        self.setTextFormat(Qt.RichText)
        self.setOpenExternalLinks(True)

class Wordless_Table_Item(QTableWidgetItem):
    def read_data(self):
        if self.column() in self.tableWidget().headers_cumulative:
            return self.val_raw
        elif self.column() in self.tableWidget().headers_num:
            return self.val
        else:
            return self.text()

    def __lt__(self, other):
        return self.read_data() < other.read_data()

class Wordless_Table(QTableWidget):
    def __init__(self, main, headers, header_orientation = 'horizontal', cols_stretch = []):
        self.main = main
        self.headers = headers
        self.header_orientation = header_orientation

        if header_orientation == 'horizontal':
            super().__init__(1, len(self.headers), self.main)

            self.setHorizontalHeaderLabels(self.headers)
        else:
            super().__init__(len(self.headers), 1, self.main)

            self.setVerticalHeaderLabels(self.headers)

        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        for col in self.find_col(cols_stretch):
            self.horizontalHeader().setSectionResizeMode(col, QHeaderView.Stretch)

        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)
        self.horizontalHeader().setHighlightSections(False)
        self.verticalHeader().setHighlightSections(False)

        self.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)

        if self.header_orientation == 'horizontal':
            self.setStyleSheet(''' 
                                   QTableView {
                                       outline: none;
                                       color: #292929;
                                   }

                                   QTableView::item:hover {
                                       background-color: #EEE;
                                       color: #292929;
                                   }
                                   QTableView::item:selected {
                                       background-color: #EEE;
                                       color: #292929;
                                   }

                                   QHeaderView::section {
                                       color: #FFF;
                                       font-weight: bold;
                                   }

                                   QHeaderView::section:horizontal {
                                       background-color: #5C88C5;
                                   }
                                   QHeaderView::section:horizontal:hover {
                                       background-color: #3265B2;
                                   }
                                   QHeaderView::section:horizontal:pressed {
                                       background-color: #3265B2;
                                   }

                                   QHeaderView::section:vertical {
                                       background-color: #737373;
                                   }
                                   QHeaderView::section:vertical:hover {
                                       background-color: #606060;
                                   }
                                   QHeaderView::section:vertical:pressed {
                                       background-color: #606060;
                                   }
                               ''')
        else:
            self.setStyleSheet('''

                                   QTableView {
                                       outline: none;
                                       color: #292929;
                                   }

                                   QTableView::item:hover {
                                       background-color: #EEE;
                                   }
                                   QTableView::item:selected {
                                       background-color: #EEE;
                                       color: #292929;
                                   }

                                   QHeaderView::section {
                                       color: #FFF;
                                       font-weight: bold;
                                   }

                                   QHeaderView::section:horizontal {
                                       background-color: #888;
                                   }
                                   QHeaderView::section:horizontal:hover {
                                       background-color: #777;
                                   }
                                   QHeaderView::section:horizontal:pressed {
                                       background-color: #666;
                                   }

                                   QHeaderView::section:vertical {
                                       background-color: #5C88C5;
                                   }
                                   QHeaderView::section:vertical:hover {
                                       background-color: #3265B2;
                                   }
                                   QHeaderView::section:vertical:pressed {
                                       background-color: #264E8C;
                                   }
                               ''')

    def insert_row(self, i, label):
        super().insertRow(i)

        self.setVerticalHeaderItem(i, QTableWidgetItem(label))

    def insert_col(self, i, label):
        super().insertColumn(i)

        self.setHorizontalHeaderItem(i, QTableWidgetItem(label))

    def clear_table(self, header_count = 1):
        self.clearContents()

        if self.header_orientation == 'horizontal':
            self.setColumnCount(len(self.headers))
            self.setRowCount(header_count)

            self.setHorizontalHeaderLabels(self.headers)
        else:
            self.setRowCount(len(self.headers))
            self.setColumnCount(header_count)

            self.setVerticalHeaderLabels(self.headers)

    def get_selected_rows(self):
        return list(set([index.row() for index in self.selectedIndexes()]))

    def find_row(self, text):
        def find(text):
            for row in range(self.rowCount()):
                if self.verticalHeaderItem(row).text() == text:
                    return row

        if type(text) == list:
            return [find(text_item) for text_item in text]
        else:
            return find(text)

    def find_rows(self, text):
        return [row
                for row in range(self.columnCount())
                if text in self.verticalHeaderItem(row).text()]

    def find_col(self, text):
        def find(text):
            for col in range(self.columnCount()):
                if self.horizontalHeaderItem(col).text() == text:
                    return col

        if type(text) == list:
            return [find(text_item) for text_item in text]
        else:
            return find(text)

    def find_cols(self, text):
        return [col
                for col in range(self.columnCount())
                if text in self.horizontalHeaderItem(col).text()]

    def find_header(self, text):
        if self.header_orientation == 'horizontal':
            return self.find_col(text)
        else:
            return self.find_row(text)

    def find_headers(self, text):
        if self.header_orientation == 'horizontal':
            return self.find_cols(text)
        else:
            return self.find_rows(text)

class Wordless_Table_Data(Wordless_Table):
    def __init__(self, main, headers, header_orientation = 'horizontal', cols_stretch = [],
                 headers_num = [], headers_pct = [], headers_cumulative = [], cols_breakdown = [],
                 sorting_enabled = False, drag_drop_enabled = False):
        super().__init__(main, headers, header_orientation, cols_stretch)

        self.headers_num_old = headers_num
        self.headers_pct_old = headers_pct
        self.headers_cumulative_old = headers_cumulative
        self.cols_breakdown_old = cols_breakdown

        self.sorting_enabled = sorting_enabled

        if sorting_enabled:
            self.setSortingEnabled(True)

            if header_orientation == 'horizontal':
                self.horizontalHeader().sortIndicatorChanged.connect(self.sorting_changed)
            else:
                self.verticalHeader().sortIndicatorChanged.connect(self.sorting_changed)

        if drag_drop_enabled:
            self.setDragEnabled(True)
            self.setAcceptDrops(True)
            self.viewport().setAcceptDrops(True)
            self.setDragDropMode(QAbstractItemView.InternalMove)
            self.setDragDropOverwriteMode(False)

        self.itemChanged.connect(self.item_changed)
        self.itemSelectionChanged.connect(self.selection_changed)

        self.button_export_selected = QPushButton(self.tr('Export Selected...'), self)
        self.button_export_all = QPushButton(self.tr('Export All...'), self)
        self.button_clear = QPushButton(self.tr('Clear'), self)

        self.button_export_selected.clicked.connect(self.export_selected)
        self.button_export_all.clicked.connect(self.export_all)
        self.button_clear.clicked.connect(lambda: self.clear_table())

        self.clear_table()

    def dropEvent(self, event):
        rows_dragged = []

        if self.indexAt(event.pos()).row() == -1:
            row_dropped = self.rowCount()
        else:
            row_dropped = self.indexAt(event.pos()).row()

        selected_rows = self.selected_rows()

        self.blockSignals(True)

        for row in selected_rows:
            rows_dragged.append([])

            for col in range(self.columnCount()):
                if self.item(row, col):
                    rows_dragged[-1].append(self.takeItem(row, col))
                else:
                    rows_dragged[-1].append(self.cellWidget(row, col))

        for i in reversed(selected_rows):
            self.removeRow(i)

            if i < row_dropped:
                row_dropped -= 1

        for row, items in enumerate(rows_dragged):
            self.insertRow(row_dropped + row)

            for col, item in enumerate(items):
                if isinstance(item, QTableWidgetItem):
                    self.setItem(row_dropped + row, col, item)

                    self.item(row, col).setSelected(True)
                elif isinstance(item, QComboBox):
                    item_combo_box = wordless_box.Wordless_Combo_Box(self.main)
                    item_combo_box.addItems([item.itemText(i) for i in range(item.count())])
                    item_combo_box.setCurrentText(item.currentText())

                    self.setCellWidget(row_dropped + row, col, item_combo_box)

        self.blockSignals(False)

        self.itemChanged.emit(self.item(0, 0))

        event.accept()

    def item_changed(self):
        rows_visible = len([i for i in range(self.rowCount()) if not self.isRowHidden(i)])

        if [i for i in range(self.columnCount()) if self.item(0, i)] and rows_visible:
            self.button_export_all.setEnabled(True)
        else:
            self.button_export_all.setEnabled(False)

        self.selection_changed()

    def selection_changed(self):
        if self.selectedIndexes() and [i for i in range(self.columnCount()) if self.item(0, i)]:
            self.button_export_selected.setEnabled(True)
        else:
            self.button_export_selected.setEnabled(False)

    def sorting_changed(self, logicalIndex, order):
        if [i for i in range(self.columnCount()) if self.item(0, i)]:
            self.update_ranks()

            if self.show_cumulative:
                self.toggle_cumulative()

            self.update_items_width()

    def insert_row(self, i, label, num = False, pct = False, cumulative = False):
        headers_num = [self.verticalHeaderItem(row).text() for row in self.headers_num]
        headers_pct = [self.verticalHeaderItem(row).text() for row in self.headers_pct]
        headers_cumulative = [self.verticalHeaderItem(row).text() for row in self.headers_cumulative]

        super().insert_row(i, label)

        if num:
            headers_num += [label]
        if pct:
            headers_pct += [label]
        if cumulative:
            headers_cumulative += [label]

        self.headers_num = set(self.find_row(headers_num))
        self.headers_pct = set(self.find_row(headers_pct))
        self.headers_cumulative = set(self.find_row(headers_cumulative))

    def insert_col(self, i, label, num = False, pct = False, cumulative = False, breakdown = False):
        if self.header_orientation == 'horizontal':
            headers_num = [self.horizontalHeaderItem(col).text() for col in self.headers_num]
            headers_pct = [self.horizontalHeaderItem(col).text() for col in self.headers_pct]
            headers_cumulative = [self.horizontalHeaderItem(col).text() for col in self.headers_cumulative]
        cols_breakdown = [self.horizontalHeaderItem(col).text() for col in self.cols_breakdown]

        super().insert_col(i, label)

        if num:
            headers_num += [label]
        if pct:
            headers_pct += [label]
        if cumulative:
            headers_cumulative += [label]
        if breakdown:
            cols_breakdown += [label]

        if self.header_orientation == 'horizontal':
            self.headers_num = set(self.find_col(headers_num))
            self.headers_pct = set(self.find_col(headers_pct))
            self.headers_cumulative = set(self.find_col(headers_cumulative))
        self.cols_breakdown = set(self.find_col(cols_breakdown))

    def set_item_num(self, row, col, item):
        item.setFont(QFont(self.main.settings_custom['general']['font_monospace']))
        item.setTextAlignment(Qt.AlignHCenter | Qt.AlignVCenter)

        super().setItem(row, col, item)

    def set_item_num_int(self, row, col, val):
        item = Wordless_Table_Item()

        item.val = int(val)

        self.set_item_num(row, col, item)

    def set_item_num_float(self, row, col, val):
        item = Wordless_Table_Item()

        item.val = float(val)

        self.set_item_num(row, col, item)

    def set_item_num_pct(self, row, col, val, total = -1):
        item = Wordless_Table_Item()

        item.val = int(val)

        if total > -1:
            item.total = int(total)

        self.set_item_num(row, col, item)

    def set_item_num_cumulative(self, row, col, val):
        item = Wordless_Table_Item()

        item.val = int(val)
        item.val_raw = item.val
        item.val_cumulative = 0

        self.set_item_num(row, col, item)

    def update_items_width(self):
        precision_val = self.main.settings_custom['general']['precision_decimal']
        precision_pct = self.main.settings_custom['general']['precision_pct']
        len_pct = precision_pct + 5

        rows_hidden = [row for row in range(self.rowCount()) if self.isRowHidden(row)]

        self.hide()
        self.blockSignals(True)

        if self.sorting_enabled:
            self.setSortingEnabled(False)

        self.setUpdatesEnabled(False)

        if self.header_orientation == 'horizontal':
            for col in self.headers_num:
                max_val = max([self.item(row, col).val
                               for row in range(self.rowCount())
                               if self.item(row, col).val != float('inf')])

                # p-value
                if self.tr('p-value') in self.horizontalHeaderItem(col).text():
                    precision_val = self.main.settings_custom['general']['precision_p_value']
                else:
                    precision_val = self.main.settings_custom['general']['precision_decimal']

                if type(max_val) == int:
                    len_val = len(f'{max_val:,}')
                else:
                    len_val = len(f'{max_val:,.{precision_val}f}')

                if col in self.headers_pct and self.show_pct:
                    pct_max = max([self.item(row, col).pct for row in range(self.rowCount())])
                    len_pct = len(f'{pct_max:.{precision_pct}%}')

                    if type(max_val) == int:
                        for row in range(self.rowCount()):
                            if not self.isRowHidden(row):
                                item = self.item(row, col)
                                
                                item.setText(f'{item.val:>{len_val},}/{item.pct:<{len_pct}.{precision_pct}%}')
                    else:
                        for row in range(self.rowCount()):
                            if not self.isRowHidden(row):
                                item = self.item(row, col)

                                if item.val == float('inf'):
                                    item.setText('+∞')
                                elif item.val == float('-inf'):
                                    item.setText('-∞')
                                else:
                                    item.setText(f'{item.val:>{len_val},.{precision_val}}/{item.pct:<{len_pct}.{precision_pct}%}')
                else:
                    if type(max_val) == int:
                        for row in range(self.rowCount()):
                            if not self.isRowHidden(row):
                                item = self.item(row, col)

                                item.setText(f'{item.val:>{len_val},}')
                    else:
                        for row in range(self.rowCount()):
                            if not self.isRowHidden(row):
                                item = self.item(row, col)

                                if item.val == float('inf'):
                                    item.setText('+ ∞')
                                elif item.val == float('-inf'):
                                    item.setText('- ∞')
                                else:
                                    item.setText(f'{item.val:>{len_val},.{precision_val}f}')
        else:
            len_vals = []

            max_vals = [max([self.item(row, col).val
                             for row in range(self.rowCount())])
                        for col in range(self.columnCount())]
            max_pcts = [max([self.item(row, col).pct
                             for row in range(self.rowCount())
                             if row in self.headers_pct])
                        for col in range(self.columnCount())]

            for max_val in max_vals:
                if type(max_val) == int:
                    len_vals.append(len(f'{max_val:,}'))
                else:
                    len_vals.append(len(f'{max_val:,.{precision_val}f}'))

            len_pcts = [len(f'{max_pct:.{precision_pct}%}') for max_pct in max_pcts]

            for row in self.headers_num:
                if row in self.headers_pct and self.show_pct:
                    if type(self.item(row, 0).val) == int:
                        for col in range(self.columnCount()):
                            if not self.isColumnHidden(col):
                                item = self.item(row, col)
                                
                                item.setText(f'{item.val:>{len_vals[col]},}/{item.pct:<{len_pcts[col]}.{precision_pct}%}')
                    else:
                        for col in range(self.columnCount()):
                            if not self.isColumnHidden(col):
                                item = self.item(row, col)

                                item.setText(f'{item.val:>{len_vals[col]},.{precision_val}}/{item.pct:<{len_pcts[col]}.{precision_pct}%}')
                else:
                    if type(self.item(row, 0).val) == int:
                        for col in range(self.columnCount()):
                            if not self.isColumnHidden(col):
                                item = self.item(row, col)

                                item.setText(f'{item.val:{len_vals[col]},}')
                    else:
                        for col in range(self.columnCount()):
                            if not self.isColumnHidden(col):
                                item = self.item(row, col)

                                item.setText(f'{item.val:{len_vals[col]},.{precision_val}f}')

        self.blockSignals(False)

        if self.sorting_enabled:
            self.setSortingEnabled(True)

        self.setUpdatesEnabled(True)
        self.show()

        self.setUpdatesEnabled(False)

        for row_hidden in rows_hidden:
            self.hideRow(row_hidden)

        self.setUpdatesEnabled(True)

    def update_ranks(self):
        data_prev = ''
        rank_prev = 1
        rank_next = 1

        sorting_section = self.horizontalHeader().sortIndicatorSection()
        col_rank = self.find_col(self.tr('Rank'))
        rows_hidden = [(self.item(row, 0).text(), self.item(row, 1).text())
                       for row in range(self.rowCount()) if self.isRowHidden(row)]

        self.sortByColumn(sorting_section, self.horizontalHeader().sortIndicatorOrder())

        rows_hidden_sorted = []

        for text0, text1 in rows_hidden:
            if (self.findItems(text0, Qt.MatchExactly)[0].column() == 0 and
                self.findItems(text1, Qt.MatchExactly)[0].column() == 1):
                rows_hidden_sorted.append(self.findItems(text0, Qt.MatchExactly)[0].row())

        for row in range(self.rowCount()):
            if row not in rows_hidden_sorted:
                data_cur = self.item(row, sorting_section).read_data()

                if data_cur == data_prev:
                    self.item(row, col_rank).val = rank_prev
                else:
                    self.item(row, col_rank).val = rank_next

                    rank_prev = rank_next

                rank_next += 1
                data_prev = data_cur

        self.setUpdatesEnabled(False)

        for row in rows_hidden_sorted:
            self.hideRow(row)

        self.setUpdatesEnabled(True)

    def toggle_pct(self):
        if self.header_orientation == 'horizontal':
            for col in self.headers_pct:
                if self.item(0, col) and not hasattr(self.item(0, col), 'total'):
                    total = sum([self.item(row, col).val_raw for row in range(self.rowCount())])

                    for row in range(self.rowCount()):
                        self.item(row, col).total = total

                for row in range(self.rowCount()):
                    item = self.item(row, col)

                    item.pct = item.val / item.total if item.total else 0
        else:
            for row in self.headers_pct:
                if self.item(row, 0) and not hasattr(self.item(row, 0), 'total'):
                    total = self.item(row, self.columnCount() - 1).val

                    for col in range(self.columnCount()):
                        self.item(row, col).total = total

                for col in range(self.columnCount()):
                    item = self.item(row, col)

                    item.pct = item.val / item.total if item.total else 0

    def toggle_cumulative(self):
        if self.header_orientation == 'horizontal':
            for col in self.headers_cumulative:
                val_cumulative = 0

                for row in range(self.rowCount()):
                    if not self.isRowHidden(row):
                        item = self.item(row, col)

                        val_cumulative += item.val
                        item.val_cumulative = val_cumulative
           
            if self.show_cumulative:
                for col in self.headers_cumulative:
                    for row in range(self.rowCount()):
                        item = self.item(row, col)

                        item.val = item.val_cumulative
                        item.pct = item.val / item.total if item.total else 0
            else:
                for col in self.headers_cumulative:
                    for row in range(self.rowCount()):
                        item = self.item(row, col)

                        item.val = item.val_raw
                        item.pct = item.val / item.total if item.total else 0
        else:
            for row in self.headers_cumulative:
                val_cumulative = 0

                for col in range(self.columnCount() - 1):
                    if not self.isColumnHidden(col):
                        item = self.item(row, col)

                        val_cumulative += item.val
                        item.val_cumulative = val_cumulative

                self.item(row, self.columnCount() - 1).val_cumulative = val_cumulative
           
            if self.show_cumulative:
                for row in self.headers_cumulative:
                    for col in range(self.columnCount()):
                        item = self.item(row, col)

                        item.val = item.val_cumulative
                        item.pct = item.val / item.total if item.total else 0
            else:
                for row in self.headers_cumulative:
                    for col in range(self.columnCount()):
                        item = self.item(row, col)

                        item.val = item.val_raw
                        item.pct = item.val / item.total if item.total else 0

    def toggle_breakdown(self):
        self.setUpdatesEnabled(False)

        for col in self.cols_breakdown:
            if self.show_breakdown:
                self.showColumn(col)
            else:
                self.hideColumn(col)

        self.setUpdatesEnabled(True)

    def filter_table(self):
        rank_prev = 1
        rank_next = 1
        data_prev = ''

        self.setUpdatesEnabled(False)
        
        for i, filters in enumerate(self.row_filters):
            if all(filters):
                self.showRow(i)
            else:
                self.hideRow(i)

        self.setUpdatesEnabled(True)

        self.toggle_cumulative()
        self.update_ranks()
        self.update_items_width()

        self.itemChanged.emit(self.item(0, 0))

    def selected_rows(self):
        return sorted(set([index.row() for index in self.selectedIndexes()]))

    def export_selected(self):
        rows_export = sorted({index.row() for index in self.selectedIndexes()})

        self.export_all(rows_export = rows_export)

    def export_all(self, rows_export = []):
        def set_cell_styles(cell, item, item_type = 'item'):
            if item_type == 'header_horizontal':
                cell.font = openpyxl.styles.Font(name = item.font().family(),
                                                 size = 8,
                                                 bold = True,
                                                 color = 'FFFFFF')

                if self.header_orientation == 'horizontal':
                    cell.fill = openpyxl.styles.PatternFill(fill_type = 'solid', fgColor = '5C88C5')
                else:
                    cell.fill = openpyxl.styles.PatternFill(fill_type = 'solid', fgColor = '888888')
            elif item_type == 'header_vertical':
                cell.font = openpyxl.styles.Font(name = item.font().family(),
                                                 size = 8,
                                                 bold = True,
                                                 color = 'FFFFFF')

                cell.fill = openpyxl.styles.PatternFill(fill_type = 'solid', fgColor = '5C88C5')
            else:
                cell.font = openpyxl.styles.Font(name = item.font().family(),
                                                 size = 8,
                                                 color = '292929')

            cell.alignment = openpyxl.styles.Alignment(horizontal = 'center',
                                                       vertical = 'center',
                                                       wrap_text = True)

        (file_path,
         file_type) = QFileDialog.getSaveFileName(self,
                                                  self.tr('Export Table'),
                                                  self.main.settings_custom['export']['tables_default_path'],
                                                  ';;'.join(self.main.settings_global['file_types']['export_tables']),
                                                  self.main.settings_custom['export']['tables_default_type'])

        if file_path:
            if file_type == self.tr('Excel Workbook (*.xlsx)'):
                workbook = openpyxl.Workbook()
                worksheet = workbook.active

                worksheet.freeze_panes = 'B2'

                dpi_horizontal = QApplication.primaryScreen().logicalDotsPerInchX()
                dpi_vertical = QApplication.primaryScreen().logicalDotsPerInchY()

                if not rows_export:
                    rows_export = list(range(self.rowCount()))

                if self.header_orientation == 'horizontal':
                    # Horizontal Headers
                    for col in range(self.columnCount()):
                        worksheet.cell(1, 1 + col).value = self.horizontalHeaderItem(col).text()

                        set_cell_styles(worksheet.cell(1, 1 + col), self.horizontalHeaderItem(col), item_type = 'header_horizontal')

                        worksheet.column_dimensions[openpyxl.utils.get_column_letter(1 + col)].width = self.horizontalHeader().sectionSize(col) / dpi_horizontal * 13 + 3

                    # Cells
                    for row_cell, row_item in enumerate(rows_export):
                        for col in range(self.columnCount()):
                            worksheet.cell(2 + row_cell, 1 + col).value = self.item(row_item, col).text()

                            set_cell_styles(worksheet.cell(2 + row_cell, 1 + col), self.item(row_item, col))
                else:
                    # Horizontal Headers
                    for col in range(self.columnCount()):
                        worksheet.cell(1, 2 + col).value = self.horizontalHeaderItem(col).text()

                        set_cell_styles(worksheet.cell(1, 2 + col), self.horizontalHeaderItem(col), item_type = 'header_horizontal')

                        worksheet.column_dimensions[openpyxl.utils.get_column_letter(2 + col)].width = self.horizontalHeader().sectionSize(col) / dpi_horizontal * 13 + 3

                    worksheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = self.verticalHeader().width() / dpi_horizontal * 13 + 3

                    # Vertical Headers
                    for row_cell, row_item in enumerate(rows_export):
                        worksheet.cell(2 + row_cell, 1).value = self.verticalHeaderItem(row_item).text()

                        set_cell_styles(worksheet.cell(2 + row_cell, 1), self.verticalHeaderItem(row_item), item_type = 'header_vertical')

                    # Cells
                    for row_cell, row_item in enumerate(rows_export):
                        for col in range(self.columnCount()):
                            worksheet.cell(2 + row_cell, 2 + col).value = self.item(row_item, col).text()

                            set_cell_styles(worksheet.cell(2 + row_cell, 2 + col), self.item(row_item, col))

                # Row Height
                worksheet.row_dimensions[1].height = self.horizontalHeader().height() / dpi_vertical * 72

                for i, _ in enumerate(worksheet.rows):
                    worksheet.row_dimensions[2 + i].height = self.verticalHeader().sectionSize(0) / dpi_vertical * 72

                # Borders
                border = openpyxl.styles.Side(border_style = 'thin', color = '292929')

                for row, _ in enumerate(worksheet.rows):
                    for col, _ in enumerate(worksheet.columns):
                        worksheet.cell(row + 1, col + 1).border = openpyxl.styles.Border(left = border,
                                                                                         right = border,
                                                                                         top = border,
                                                                                         bottom = border)

                workbook.save(file_path)
            elif file_type == self.tr('CSV (Comma Delimited) (*.csv)'):
                encoding = self.main.settings_custom['export']['tables_default_encoding']

                with open(file_path, 'w', encoding = encoding, newline = '') as f:
                    csv_writer = csv.writer(f)

                    if not rows_export:
                        rows_export = list(range(self.rowCount()))

                    if self.header_orientation == 'horizontal':
                        # Horizontal Headers
                        csv_writer.writerow([self.horizontalHeaderItem(col).text().strip() for col in range(self.columnCount())])

                        # Cells
                        for row in rows_export:
                            csv_writer.writerow([self.item(row, col).text().strip() for col in range(self.columnCount())])

                    else:
                        # Horizontal Headers
                        csv_writer.writerow([''] +
                                            [self.horizontalHeaderItem(col).text().strip() for col in range(self.columnCount())])

                        # Vertical Headers & Cells
                        for row in rows_export:
                            csv_writer.writerow([self.verticalHeaderItem(row).text().strip()] +
                                                [self.item(row, col).text().strip() for col in range(self.columnCount())])

            self.main.settings_custom['export']['tables_default_path'] = os.path.split(file_path)[0]
            self.main.settings_custom['export']['tables_default_type'] = file_type

            wordless_message_box.wordless_message_box_export_completed_table(self.main, file_path)

    def clear_table(self, header_count = 1):
        self.clearContents()

        if self.header_orientation == 'horizontal':
            self.horizontalHeader().blockSignals(True)

            self.setColumnCount(len(self.headers))
            self.setRowCount(header_count)

            self.setHorizontalHeaderLabels(self.headers)

            self.horizontalHeader().blockSignals(False)

            self.horizontalHeader().sectionCountChanged.emit(0, header_count)
        else:
            self.verticalHeader().blockSignals(True)

            self.setRowCount(len(self.headers))
            self.setColumnCount(header_count)

            self.setVerticalHeaderLabels(self.headers)

            self.verticalHeader().blockSignals(False)

            self.verticalHeader().sectionCountChanged.emit(0, header_count)

        for i in range(self.rowCount()):
            self.showRow(i)
        for i in range(self.columnCount()):
            self.showColumn(i)

        self.headers_num = set(self.find_header(self.headers_num_old))
        self.headers_pct = set(self.find_header(self.headers_pct_old))
        self.headers_cumulative = set(self.find_header(self.headers_cumulative_old))
        self.cols_breakdown = set(self.find_col(self.cols_breakdown_old))
        self.settings = self.main.settings_default

        self.item_changed()

class Wordless_Table_Data_Search(Wordless_Table_Data):
    def __init__(self, main, headers, header_orientation = 'horizontal', cols_stretch = [],
                 headers_num = [], headers_pct = [], headers_cumulative = [], cols_breakdown = [],
                 sorting_enabled = False, drag_drop_enabled = False):
        super().__init__(main, headers, header_orientation, cols_stretch,
                         headers_num, headers_pct, headers_cumulative, cols_breakdown,
                         sorting_enabled, drag_drop_enabled)

        self.label_number_results = QLabel()
        self.button_search_results = QPushButton(self.tr('Search in Results'), self)

        self.itemChanged.connect(self.results_changed)

        self.results_changed()

    def results_changed(self):
        rows_visible = len([i for i in range(self.rowCount()) if not self.isRowHidden(i)])

        if [i for i in range(self.columnCount()) if self.item(0, i)] and rows_visible:
            self.label_number_results.setText(self.tr(f'Number of Results: {rows_visible}'))

            self.button_search_results.setEnabled(True)
        else:
            self.label_number_results.setText(self.tr('Number of Results: 0'))

            self.button_search_results.setEnabled(False)
