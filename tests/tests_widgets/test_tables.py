# ----------------------------------------------------------------------
# Tests: Widgets - Tables
# Copyright (C) 2018-2025  Ye Lei (叶磊)
#
# This program is free software: you can redistribute it and/or modify
# it under the terms of the GNU General Public License as published by
# the Free Software Foundation, either version 3 of the License, or
# (at your option) any later version.
#
# This program is distributed in the hope that it will be useful,
# but WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
# GNU General Public License for more details.
#
# You should have received a copy of the GNU General Public License
# along with this program.  If not, see <https://www.gnu.org/licenses/>.
# ----------------------------------------------------------------------

import docx
import openpyxl
from PyQt5 import QtCore
from PyQt5 import QtGui

from tests import wl_test_init
from wordless.wl_dialogs import wl_dialogs_misc
from wordless.wl_widgets import wl_tables

main = wl_test_init.Wl_Test_Main()

def test_wl_table():
    table_hor = wl_tables.Wl_Table(
        main,
        headers = ('test',), header_orientation = 'hor',
        editable = True,
        drag_drop = True
    )
    table_vert = wl_tables.Wl_Table(
        main,
        headers = ('test',), header_orientation = 'vert',
        editable = False,
        drag_drop = True
    )

    table_hor.item_changed()
    table_hor.selection_changed()
    table_hor.disable_updates()
    table_hor.enable_updates()

    table_hor.is_empty()
    table_vert.is_empty()

    table_hor.is_visible()
    table_hor.is_selected()
    table_hor.get_header_labels_hor()
    table_hor.get_header_labels_vert()

    table_hor.find_header_hor('test')
    table_vert.find_header_vert('test')
    table_hor.find_headers_hor('test')
    table_vert.find_headers_vert('test')

    table_hor.clr_table()
    table_hor.get_visible_rows()
    table_hor.model().setItem(0, 0, QtGui.QStandardItem('test'))
    table_hor.get_visible_rows()

    table_hor.get_selected_rows(visible_only = True)
    table_hor.get_selected_rows(visible_only = False)

    table_hor.defaults_row = ('test',)
    table_hor.clr_table()
    table_hor._add_row()
    table_hor._add_row(row = 0)
    table_hor.add_row()
    table_hor.ins_row()
    table_hor.del_row()

    table_hor.clr_table()
    table_vert.clr_table()

    table_hor.update_gui_exp('', '')

def test_wl_worker_exp_table():
    table_hor = wl_test_init.Wl_Test_Table(
        main,
        headers = ('test',), header_orientation = 'hor'
    )
    table_vert = wl_test_init.Wl_Test_Table(
        main,
        headers = ('test',), header_orientation = 'vert'
    )

    worker_hor = wl_tables.Wl_Worker_Exp_Table(
        main,
        dialog_progress = wl_dialogs_misc.Wl_Dialog_Progress(main, ''),
        table = table_hor
    )
    worker_vert = wl_tables.Wl_Worker_Exp_Table(
        main,
        dialog_progress = wl_dialogs_misc.Wl_Dialog_Progress(main, ''),
        table = table_vert
    )

    worker_hor.clean_text_csv(['test'])
    worker_hor.remove_invalid_xml_chars('test')

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worker_hor.style_header(worksheet.cell(1, 1))

    worker_hor.style_header_hor(worksheet.cell(1, 1))
    worker_hor.style_header_vert(worksheet.cell(1, 1))
    worker_vert.style_header_vert(worksheet.cell(1, 1))
    worker_vert.style_header_hor(worksheet.cell(1, 1))

    table_hor.set_item(0, 0, 'test')
    worker_hor.style_cell_alignment(worksheet.cell(1, 1), table_hor.model().item(0, 0))
    table_hor.model().item(0, 0).setTextAlignment(QtCore.Qt.AlignLeft | QtCore.Qt.AlignTop)
    worker_hor.style_cell_alignment(worksheet.cell(1, 1), table_hor.model().item(0, 0))
    table_hor.model().item(0, 0).setTextAlignment(QtCore.Qt.AlignHCenter | QtCore.Qt.AlignVCenter)
    worker_hor.style_cell_alignment(worksheet.cell(1, 1), table_hor.model().item(0, 0))
    table_hor.model().item(0, 0).setTextAlignment(QtCore.Qt.AlignRight | QtCore.Qt.AlignBottom)
    worker_hor.style_cell_alignment(worksheet.cell(1, 1), table_hor.model().item(0, 0))
    table_hor.model().item(0, 0).setTextAlignment(QtCore.Qt.AlignJustify | QtCore.Qt.AlignBaseline)
    worker_hor.style_cell_alignment(worksheet.cell(1, 1), table_hor.model().item(0, 0))
    table_hor.set_label(0, 1, 'test')
    worker_hor.style_cell_alignment(worksheet.cell(1, 1), table_hor.indexWidget(table_hor.model().index(0, 1)))

    table_hor.headers_p_val = {0}
    table_hor.settings['tables']['precision_settings']['precision_p_vals'] = 1
    worker_hor.style_cell(worksheet.cell(1, 1), table_hor.model().item(0, 0))
    table_hor.settings['tables']['precision_settings']['precision_p_vals'] = 0
    worker_hor.style_cell(worksheet.cell(1, 1), table_hor.model().item(0, 0))

    table_hor.headers_pct = {0}
    table_hor.settings['tables']['precision_settings']['precision_pcts'] = 1
    worker_hor.style_cell(worksheet.cell(1, 1), table_hor.model().item(0, 0))
    table_hor.settings['tables']['precision_settings']['precision_pcts'] = 0
    worker_hor.style_cell(worksheet.cell(1, 1), table_hor.model().item(0, 0))

    table_hor.headers_float = {0}
    table_hor.settings['tables']['misc_settings']['show_thousand_separators'] = True
    table_hor.settings['tables']['precision_settings']['precision_decimals'] = 1
    worker_hor.style_cell(worksheet.cell(1, 1), table_hor.model().item(0, 0))
    table_hor.settings['tables']['misc_settings']['show_thousand_separators'] = False
    worker_hor.style_cell(worksheet.cell(1, 1), table_hor.model().item(0, 0))

    table_hor.headers_int = {0}
    table_hor.settings['tables']['misc_settings']['show_thousand_separators'] = True
    worker_hor.style_cell(worksheet.cell(1, 1), table_hor.model().item(0, 0))
    table_hor.settings['tables']['misc_settings']['show_thousand_separators'] = False
    worker_hor.style_cell(worksheet.cell(1, 1), table_hor.model().item(0, 0))

    table_vert.set_item(0, 0, 'test')
    worker_vert.style_cell(worksheet.cell(1, 1), table_vert.model().item(0, 0))

    worksheet.cell(1, 1).value = '<span style="color: #000;">test</span><span style="">test</span><span>test</span>test'
    worker_hor.style_cell_rich_text(worksheet.cell(1, 1), table_hor.model().item(0, 0))

    doc = docx.Document()
    para = worker_hor.add_para(doc)
    worker_hor.style_para_rich_text(
        para,
        para_text = '<span style="color: #000;">test</span><span style="">test</span><span>test</span>test',
        item = table_hor.model().item(0, 0)
    )
    worker_hor.style_para_spacing(para)

def test_wl_table_add_ins_del_clr():
    table = wl_tables.Wl_Table_Add_Ins_Del_Clr(main, headers = ('test',), col_edit = 0)
    table.defaults_row = ('test',)

    table.clr_table()
    table.item_changed()
    table.add_row()
    table.item_changed()

    table.clearSelection()
    table.selection_changed()
    table.selectAll()
    table.selection_changed()

    table.add_row()
    table.ins_row()

def test_wl_table_item():
    table_item_0 = wl_tables.Wl_Table_Item()
    table_item_1 = wl_tables.Wl_Table_Item()
    table_item_2 = wl_tables.Wl_Table_Item()

    table = wl_test_init.Wl_Test_Table(main, tab = 'test', headers = ('test',), headers_int = {'test'})
    table.model().setItem(0, 0, table_item_0)
    table.model().item(0, 0).val = 0
    table.model().setItem(0, 1, table_item_1)
    table.model().setItem(1, 0, table_item_2)
    table.model().item(1, 0).val = 1

    table_item_0.read_data()
    table_item_1.read_data()

    assert table_item_0 < table_item_2

def test_wl_table_item_err():
    table_item = wl_tables.Wl_Table_Item_Err('a')
    table_item.read_data()

    assert wl_tables.Wl_Table_Item_Err() < table_item

def test_wl_table_data():
    table_hor = wl_tables.Wl_Table_Data(
        main, tab = 'concordancer',
        headers = ('Rank',), header_orientation = 'hor',
        enable_sorting = True, generate_fig = False,
        results_search = True, results_filter = True, results_sample = True, results_sort = True
    )
    table_vert = wl_tables.Wl_Table_Data(
        main, tab = 'concordancer',
        headers = ('test',), header_orientation = 'vert',
        enable_sorting = True
    )

    table_hor.clr_table()
    table_hor.item_changed()
    table_hor.set_item_err(0, 0, 'test')
    table_hor.item_changed()

    table_hor.clr_table()
    table_hor.selection_changed()
    table_hor.set_item_err(0, 0, 'test')
    table_hor.selectAll()
    table_hor.selection_changed()

    table_hor.horizontalHeader().setSortIndicator(0, QtCore.Qt.DescendingOrder)
    table_hor.table_settings['show_cum_data'] = True
    table_hor.sorting_changed()

    main.settings_custom['file_area']['files_open'].clear()
    table_hor.file_changed()
    main.settings_custom['file_area']['files_open'] = [{'selected': True}]
    table_hor.file_changed()

    table_hor.add_header_hor('test')
    table_vert.add_header_vert('test')

    table_hor.ins_header_hor(
        0, 'test',
        is_int = True, is_float = True,
        is_pct = True, is_p_val = True,
        is_cum = True,
        is_breakdown_span_position = True,
        is_breakdown_file = True, is_total = True
    )
    table_vert.ins_header_vert(
        0, 'test',
        is_int = True, is_float = True,
        is_pct = True, is_p_val = True,
        is_cum = True
    )

    table_hor.settings['tables']['misc_settings']['show_thousand_separators'] = True
    table_hor.format_int(0)
    table_hor.settings['tables']['misc_settings']['show_thousand_separators'] = False
    table_hor.format_int(0)

    table_hor.settings['tables']['misc_settings']['show_thousand_separators'] = True
    table_hor.format_float(0, 0)
    table_hor.settings['tables']['misc_settings']['show_thousand_separators'] = False
    table_hor.format_float(0, 0)

    table_hor.format_pct(0, 0)

    table_hor.set_item_num(0, 0, 0)
    table_hor.headers_int.clear()
    table_hor.set_item_num(0, 0, 0)
    table_hor.headers_float.clear()
    table_hor.set_item_num(0, 0, 0, total = -1)
    table_hor.set_item_num(0, 0, 0, total = 0)
    table_hor.set_item_num(0, 0, 0, total = 1)
    table_hor.headers_pct.clear()
    table_hor.set_item_num(0, 0, 0)
    table_vert.set_item_num(0, 0, 0)

    table_hor.headers_pct = {0}
    table_hor.set_item_num_val(0, 0, 0)
    table_hor.headers_int = {0}
    table_hor.set_item_num_val(0, 0, 0)
    table_vert.set_item_num_val(0, 0, 0)

    table_hor.set_item_err(0, 0, 'test', alignment_hor = 'center')
    table_hor.set_item_err(0, 0, 'test', alignment_hor = 'left')
    table_hor.set_item_err(0, 0, 'test', alignment_hor = 'right')

if __name__ == '__main__':
    test_wl_table()
    test_wl_worker_exp_table()
    test_wl_table_add_ins_del_clr()
    test_wl_table_item()
    test_wl_table_item_err()
    test_wl_table_data()
