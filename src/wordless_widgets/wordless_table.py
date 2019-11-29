#
# Wordless: Widgets - Table
#
# Copyright (C) 2018-2019  Ye Lei (叶磊)
#
# This source file is licensed under GNU GPLv3.
# For details, see: https://github.com/BLKSerene/Wordless/blob/master/LICENSE.txt
#
# All other rights reserved.
#

import copy
import csv
import os
import re
import time

from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *

import bs4
import openpyxl

from wordless_checking import wordless_checking_misc
from wordless_dialogs import wordless_dialog_misc, wordless_msg_box
from wordless_text import wordless_text_processing, wordless_text_utils
from wordless_utils import wordless_misc, wordless_threading
from wordless_widgets import (wordless_box, wordless_button, wordless_label,
                              wordless_msg)

class Wordless_Worker_Export_Table(wordless_threading.Wordless_Worker):
    worker_done = pyqtSignal(bool, str)

    def run(self):
        if 'headers_int' not in self.table.__dict__:
            self.table.headers_int = []
        if 'headers_float' not in self.table.__dict__:
            self.table.headers_float = []
        if 'headers_pct' not in self.table.__dict__:
            self.table.headers_pct = []

        # Check file permissions
        try:
            if not self.rows_export:
                self.rows_export = list(range(self.table.rowCount()))

            len_rows = len(self.rows_export)

            # Excel workbooks
            if self.file_type == self.tr('Excel Workbook (*.xlsx)'):
                workbook = openpyxl.Workbook()
                worksheet = workbook.active

                # Freeze panes
                if self.table.name in ['error', 'concordancer']:
                    worksheet.freeze_panes = 'A2'
                else:
                    worksheet.freeze_panes = 'B2'

                dpi_horizontal = QApplication.primaryScreen().logicalDotsPerInchX()
                dpi_vertical = QApplication.primaryScreen().logicalDotsPerInchY()

                # Concordancer
                if self.table.name == 'concordancer':
                    # Horizontal Headers
                    for col in range(self.table.columnCount()):
                        cell = worksheet.cell(1, 1 + col)
                        cell.value = self.table.horizontalHeaderItem(col).text()

                        self.style_header_horizontal(cell, self.table.horizontalHeaderItem(col))

                        worksheet.column_dimensions[openpyxl.utils.get_column_letter(1 + col)].width = self.table.horizontalHeader().sectionSize(col) / dpi_horizontal * 13 + 3

                    # Cells
                    for row_cell, row_item in enumerate(self.rows_export):
                        for col in range(self.table.columnCount()):
                            # Left
                            if col == 0:
                                cell = worksheet.cell(2 + row_cell, 1 + col)

                                cell_val = wordless_text_utils.html_to_text(self.table.cellWidget(row_item, col).text())
                                # Remove illegal characters
                                cell_val = re.sub(openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE, '', cell_val)
                                cell.value = cell_val

                                cell.font = openpyxl.styles.Font(
                                    name = self.table.cellWidget(row_item, col).font().family(),
                                    size = 8,
                                    color = '292929'
                                )
                                cell.alignment = openpyxl.styles.Alignment(
                                    horizontal = 'right',
                                    vertical = 'center'
                                )
                            # Node
                            elif col == 1:
                                cell = worksheet.cell(2 + row_cell, 1 + col)

                                cell_val = wordless_text_utils.html_to_text(self.table.cellWidget(row_item, col).text())
                                # Remove illegal characters
                                cell_val = re.sub(openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE, '', cell_val)
                                cell.value = cell_val

                                self.style_cell_text(cell, self.table.cellWidget(row_item, col))

                                cell.font = openpyxl.styles.Font(
                                    name = self.table.cellWidget(row_item, col).font().family(),
                                    size = 8,
                                    bold = True,
                                    color = 'FF0000'
                                )
                                cell.alignment = openpyxl.styles.Alignment(
                                    horizontal = 'center',
                                    vertical = 'center'
                                )
                            # Right
                            elif col == 2:
                                cell = worksheet.cell(2 + row_cell, 1 + col)

                                cell_val = wordless_text_utils.html_to_text(self.table.cellWidget(row_item, col).text())
                                # Remove illegal characters
                                cell_val = re.sub(openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE, '', cell_val)
                                cell.value = cell_val

                                cell.font = openpyxl.styles.Font(
                                    name = self.table.cellWidget(row_item, col).font().family(),
                                    size = 8,
                                    color = '292929'
                                )
                                cell.alignment = openpyxl.styles.Alignment(
                                    horizontal = 'left',
                                    vertical = 'center'
                                )
                            else:
                                cell = worksheet.cell(2 + row_cell, 1 + col)

                                cell_val = cell_text = self.table.item(row_item, col).text()
                                # Remove illegal characters
                                cell_val = re.sub(openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE, '', cell_val)
                                cell.value = cell_val

                                if (col in self.table.headers_int or
                                    col in self.table.headers_float or
                                    col in self.table.headers_pct):
                                    self.style_cell_num(cell, self.table.item(row_item, col))
                                else:
                                    self.style_cell_text(cell, self.table.item(row_item, col))

                            self.progress_updated.emit(self.tr(f'Exporting table ... ({row_cell + 1} / {len_rows})'))
                else:
                    if self.table.header_orientation == 'horizontal':
                        # Horizontal Headers
                        for col in range(self.table.columnCount()):
                            cell = worksheet.cell(1, 1 + col)
                            cell.value = self.table.horizontalHeaderItem(col).text()

                            self.style_header_horizontal(cell, self.table.horizontalHeaderItem(col))

                            worksheet.column_dimensions[openpyxl.utils.get_column_letter(1 + col)].width = self.table.horizontalHeader().sectionSize(col) / dpi_horizontal * 13 + 3

                        # Cells
                        for row_cell, row_item in enumerate(self.rows_export):
                            for col in range(self.table.columnCount()):
                                cell = worksheet.cell(2 + row_cell, 1 + col)

                                cell_val = self.table.item(row_item, col).text()
                                # Remove illegal characters
                                cell_val = re.sub(openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE, '', cell_val)
                                cell.value = cell_val

                                if (col in self.table.headers_int or
                                    col in self.table.headers_float or
                                    col in self.table.headers_pct):
                                    self.style_cell_num(cell, self.table.item(row_item, col))
                                else:
                                    self.style_cell_text(cell, self.table.item(row_item, col))

                            self.progress_updated.emit(self.tr(f'Exporting table ... ({row_cell + 1} / {len_rows})'))
                    else:
                        # Horizontal Headers
                        for col in range(self.table.columnCount()):
                            cell = worksheet.cell(1, 2 + col) 
                            cell.value = self.table.horizontalHeaderItem(col).text()

                            self.style_header_horizontal(cell, self.table.horizontalHeaderItem(col))

                            worksheet.column_dimensions[openpyxl.utils.get_column_letter(2 + col)].width = self.table.horizontalHeader().sectionSize(col) / dpi_horizontal * 13 + 3

                        worksheet.column_dimensions[openpyxl.utils.get_column_letter(1)].width = self.table.verticalHeader().width() / dpi_horizontal * 13 + 3

                        # Vertical Headers
                        for row_cell, row_item in enumerate(self.rows_export):
                            cell = worksheet.cell(2 + row_cell, 1)
                            cell.value = self.table.verticalHeaderItem(row_item).text()

                            self.style_header_vertical(cell, self.table.verticalHeaderItem(row_item))

                        # Cells
                        for row_cell, row_item in enumerate(self.rows_export):
                            for col in range(self.table.columnCount()):
                                cell = worksheet.cell(2 + row_cell, 2 + col)

                                cell_val = self.table.item(row_item, col).text()
                                # Remove illegal characters
                                cell_val = re.sub(openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE, '', cell_val)
                                cell.value = cell_val

                                if (col in self.table.headers_int or
                                    col in self.table.headers_float or
                                    col in self.table.headers_pct):
                                    self.style_cell_num(cell, self.table.item(row_item, col))
                                else:
                                    self.style_cell_text(cell, self.table.item(row_item, col))

                            self.progress_updated.emit(self.tr(f'Exporting table ... ({row_cell + 1} / {len_rows})'))

                # Row Height
                worksheet.row_dimensions[1].height = self.table.horizontalHeader().height() / dpi_vertical * 72

                for i, _ in enumerate(worksheet.rows):
                    worksheet.row_dimensions[2 + i].height = self.table.verticalHeader().sectionSize(0) / dpi_vertical * 72

                self.progress_updated.emit(self.tr(f'Saving file ...'))

                workbook.save(self.file_path)
            # CSV files
            elif self.file_type == self.tr('CSV File (*.csv)'):
                encoding = self.main.settings_custom['export']['tables']['default_encoding']

                with open(self.file_path, 'w', encoding = encoding, newline = '') as f:
                    csv_writer = csv.writer(f)

                    # Concordancer
                    if self.table.name == 'concordancer':
                        # Horizontal Headers
                        csv_writer.writerow([self.table.horizontalHeaderItem(col).text().strip()
                                             for col in range(self.table.columnCount())])

                        # Cells
                        for i, row in enumerate(self.rows_export):
                            row_to_export = []

                            for col in range(self.table.columnCount()):
                                if self.table.item(row, col):
                                    cell_text = self.table.item(row, col).text()
                                else:
                                    cell_text = self.table.cellWidget(row, col).text()
                                    cell_text = wordless_text_utils.html_to_text(cell_text)

                                row_to_export.append(cell_text)

                            csv_writer.writerow(row_to_export)

                            self.progress_updated.emit(self.tr(f'Exporting table ... ({i + 1} / {len_rows})'))
                    else:
                        if self.table.header_orientation == 'horizontal':
                            # Horizontal Headers
                            csv_writer.writerow([self.table.horizontalHeaderItem(col).text().strip()
                                                 for col in range(self.table.columnCount())])

                            # Cells
                            for i, row in enumerate(self.rows_export):
                                row_to_export = []

                                for col in range(self.table.columnCount()):
                                    row_to_export.append(self.table.item(row, col).text().strip())

                                csv_writer.writerow(row_to_export)

                                self.progress_updated.emit(self.tr(f'Exporting table ... ({i + 1} / {len_rows})'))
                        else:
                            # Horizontal Headers
                            csv_writer.writerow([''] +
                                                [self.table.horizontalHeaderItem(col).text().strip()
                                                 for col in range(self.table.columnCount())])

                            # Vertical Headers & Cells
                            for i, row in enumerate(self.rows_export):
                                row_to_export = [self.table.verticalHeaderItem(row).text().strip()]

                                for col in range(self.table.columnCount()):
                                    row_to_export.append(self.table.item(row, col).text().strip())

                                csv_writer.writerow(row_to_export)

                                self.progress_updated.emit(self.tr(f'Exporting table ... ({i + 1} / {len_rows})'))

            self.main.settings_custom['export']['tables']['default_path'] = wordless_misc.get_normalized_dir(self.file_path)
            self.main.settings_custom['export']['tables']['default_type'] = self.file_type

            export_success = True
        except PermissionError:
            export_success = False

        self.worker_done.emit(export_success, self.file_path)

    def remove_illegal_chars(self, text):
        return re.sub(openpyxl.cell.cell.ILLEGAL_CHARACTERS_RE, '', text)

    def style_header_horizontal(self, cell, item):
        cell.font = openpyxl.styles.Font(
            name = item.font().family(),
            size = 8,
            bold = True,
            color = 'FFFFFF'
        )

        if self.table.header_orientation == 'horizontal':
            cell.fill = openpyxl.styles.PatternFill(
                fill_type = 'solid',
                fgColor = '5C88C5'
            )
        else:
            cell.fill = openpyxl.styles.PatternFill(
                fill_type = 'solid',
                fgColor = '888888'
            )

        cell.alignment = openpyxl.styles.Alignment(
            horizontal = 'center',
            vertical = 'center',
            wrap_text = True
        )

    def style_header_vertical(self, cell, item):
        cell.font = openpyxl.styles.Font(
            name = item.font().family(),
            size = 8,
            bold = True,
            color = 'FFFFFF'
        )

        if self.table.header_orientation == 'horizontal':
            cell.fill = openpyxl.styles.PatternFill(
                fill_type = 'solid',
                fgColor = '888888'
            )

            cell.alignment = openpyxl.styles.Alignment(
                horizontal = 'right',
                vertical = 'center',
                wrap_text = True
          )
        else:
            cell.fill = openpyxl.styles.PatternFill(
                fill_type = 'solid',
                fgColor = '5C88C5'
            )

            cell.alignment = openpyxl.styles.Alignment(
                horizontal = 'left',
                vertical = 'center',
                wrap_text = True
            )

    def style_cell_text(self, cell, item):
        cell.font = openpyxl.styles.Font(
            name = item.font().family(),
            size = 8,
            color = '292929'
        )

        cell.alignment = openpyxl.styles.Alignment(
            horizontal = 'left',
            vertical = 'center',
            wrap_text = True
        )

    def style_cell_num(self, cell, item):
        cell.font = openpyxl.styles.Font(
            name = item.font().family(),
            size = 8,
            color = '292929'
        )

        cell.alignment = openpyxl.styles.Alignment(
            horizontal = 'right',
            vertical = 'center',
            wrap_text = True
        )

class Wordless_Table_Item(QTableWidgetItem):
    def read_data(self):
        if (self.column() in self.tableWidget().headers_int or
            self.column() in self.tableWidget().headers_float or
            self.column() in self.tableWidget().headers_pct):
            return self.val
        else:
            return self.text()

    def __lt__(self, other):
        return self.read_data() < other.read_data()

class Wordless_Table(QTableWidget):
    def __init__(self, parent, headers, header_orientation = 'horizontal',
                 cols_stretch = [], drag_drop_enabled = False):
        self.main = wordless_misc.find_wordless_main(parent)
        self.headers = headers
        self.header_orientation = header_orientation
        self.cols_stretch = cols_stretch
        self.settings = self.main.settings_custom

        if header_orientation == 'horizontal':
            super().__init__(1, len(self.headers), parent)

            self.setHorizontalHeaderLabels(self.headers)
        else:
            super().__init__(len(self.headers), 1, parent)

            self.setVerticalHeaderLabels(self.headers)

        self.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
        self.verticalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)

        for col in self.find_col(cols_stretch):
            self.horizontalHeader().setSectionResizeMode(col, QHeaderView.Stretch)

        self.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.setSelectionBehavior(QAbstractItemView.SelectRows)

        if drag_drop_enabled:
            self.setDragEnabled(True)
            self.setAcceptDrops(True)
            self.viewport().setAcceptDrops(True)
            self.setDragDropMode(QAbstractItemView.InternalMove)
            self.setDragDropOverwriteMode(False)

        self.horizontalHeader().setHighlightSections(False)
        self.verticalHeader().setHighlightSections(False)

        self.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)

        if self.header_orientation == 'horizontal':
            self.setStyleSheet(''' 
                                   QTableView {
                                       outline: none;
                                       color: #292929;
                                   }

                                   QTableView::item:hover {
                                       background-color: #EEE;
                                       color: #292929;
                                   }
                                   QTableView::item:selected {
                                       background-color: #EEE;
                                       color: #292929;
                                   }

                                   QHeaderView::section {
                                       color: #FFF;
                                       font-weight: bold;
                                   }

                                   QHeaderView::section:horizontal {
                                       background-color: #5C88C5;
                                   }
                                   QHeaderView::section:horizontal:hover {
                                       background-color: #3265B2;
                                   }
                                   QHeaderView::section:horizontal:pressed {
                                       background-color: #3265B2;
                                   }

                                   QHeaderView::section:vertical {
                                       background-color: #737373;
                                   }
                                   QHeaderView::section:vertical:hover {
                                       background-color: #606060;
                                   }
                                   QHeaderView::section:vertical:pressed {
                                       background-color: #606060;
                                   }
                               ''')
        else:
            self.setStyleSheet('''

                                   QTableView {
                                       outline: none;
                                       color: #292929;
                                   }

                                   QTableView::item:hover {
                                       background-color: #EEE;
                                   }
                                   QTableView::item:selected {
                                       background-color: #EEE;
                                       color: #292929;
                                   }

                                   QHeaderView::section {
                                       color: #FFF;
                                       font-weight: bold;
                                   }

                                   QHeaderView::section:horizontal {
                                       background-color: #888;
                                   }
                                   QHeaderView::section:horizontal:hover {
                                       background-color: #777;
                                   }
                                   QHeaderView::section:horizontal:pressed {
                                       background-color: #666;
                                   }

                                   QHeaderView::section:vertical {
                                       background-color: #5C88C5;
                                   }
                                   QHeaderView::section:vertical:hover {
                                       background-color: #3265B2;
                                   }
                                   QHeaderView::section:vertical:pressed {
                                       background-color: #264E8C;
                                   }
                               ''')

        self.itemChanged.connect(self.item_changed)

    def dropEvent(self, event):
        rows_dragged = []

        if self.indexAt(event.pos()).row() == -1:
            row_dropped = self.rowCount()
        else:
            row_dropped = self.indexAt(event.pos()).row()

        selected_rows = self.get_selected_rows()

        self.blockSignals(True)

        for row in selected_rows:
            rows_dragged.append([])

            for col in range(self.columnCount()):
                if self.item(row, col):
                    rows_dragged[-1].append(self.takeItem(row, col))
                else:
                    rows_dragged[-1].append(self.cellWidget(row, col))

        for i in reversed(selected_rows):
            self.removeRow(i)

            if i < row_dropped:
                row_dropped -= 1

        for row, items in enumerate(rows_dragged):
            self.insertRow(row_dropped + row)

            for col, item in enumerate(items):
                if isinstance(item, QTableWidgetItem):
                    self.setItem(row_dropped + row, col, item)

                    self.item(row, col).setSelected(True)
                elif isinstance(item, QComboBox):
                    item_combo_box = wordless_box.Wordless_Combo_Box(self)
                    item_combo_box.addItems([item.itemText(i) for i in range(item.count())])
                    item_combo_box.setCurrentText(item.currentText())

                    self.setCellWidget(row_dropped + row, col, item_combo_box)

        self.blockSignals(False)

        self.itemChanged.emit(self.item(0, 0))

        event.accept()

    def item_changed(self):
        cols_stretch = self.find_col(self.cols_stretch)

        self.resizeRowsToContents()

        for i in range(self.columnCount()):
            if i not in cols_stretch:
                self.resizeColumnToContents(i)

    def insert_row(self, i, label):
        super().insertRow(i)

        self.setVerticalHeaderItem(i, QTableWidgetItem(label))

    def insert_col(self, i, label):
        super().insertColumn(i)

        self.setHorizontalHeaderItem(i, QTableWidgetItem(label))

    def export_selected(self):
        rows_export = sorted({index.row() for index in self.selectedIndexes()})

        self.export_all(rows_export = rows_export)

    def export_all(self, rows_export = []):
        def update_gui(export_success, file_path):
            if export_success:
                wordless_msg_box.wordless_msg_box_export_table_success(self.main, file_path)
            else:
                wordless_msg_box.wordless_msg_box_export_table_error(self.main, file_path)

        default_dir = self.main.settings_custom['export']['tables']['default_path']

        (file_path,
         file_type) = QFileDialog.getSaveFileName(self,
                                                  self.tr('Export Table'),
                                                  wordless_checking_misc.check_dir(default_dir),
                                                  ';;'.join(self.main.settings_global['file_types']['export_tables']),
                                                  self.main.settings_custom['export']['tables']['default_type'])

        if file_path:
            dialog_progress = wordless_dialog_misc.Wordless_Dialog_Progress_Export_Table(self.main)

            worker_export_table = Wordless_Worker_Export_Table(
                self.main,
                dialog_progress = dialog_progress,
                update_gui = update_gui,
                table = self,
                file_path = file_path,
                file_type = file_type,
                rows_export = rows_export
            )

            thread_export_table = wordless_threading.Wordless_Thread(worker_export_table)
            thread_export_table.start_worker()

    def clear_table(self, header_count = 1):
        self.clearContents()

        if self.header_orientation == 'horizontal':
            self.setColumnCount(len(self.headers))
            self.setRowCount(header_count)

            self.setHorizontalHeaderLabels(self.headers)
        else:
            self.setRowCount(len(self.headers))
            self.setColumnCount(header_count)

            self.setVerticalHeaderLabels(self.headers)

    def get_selected_rows(self):
        return sorted(set([index.row() for index in self.selectedIndexes()]))

    def find_row(self, text):
        def find(text):
            for row in range(self.rowCount()):
                if self.verticalHeaderItem(row).text() == text:
                    return row

        if type(text) == list:
            return [find(text_item) for text_item in text]
        else:
            return find(text)

    def find_rows(self, text):
        return [row
                for row in range(self.columnCount())
                if text in self.verticalHeaderItem(row).text()]

    def find_col(self, text):
        def find(text):
            for col in range(self.columnCount()):
                if self.horizontalHeaderItem(col).text() == text:
                    return col

        if type(text) == list:
            return [find(text_item) for text_item in text]
        else:
            return find(text)

    def find_cols(self, text):
        return [col
                for col in range(self.columnCount())
                if text in self.horizontalHeaderItem(col).text()]

    def find_header(self, text):
        if self.header_orientation == 'horizontal':
            return self.find_col(text)
        else:
            return self.find_row(text)

    def find_headers(self, text):
        if self.header_orientation == 'horizontal':
            return self.find_cols(text)
        else:
            return self.find_rows(text)

class Wordless_Table_Error(Wordless_Table):
    def __init__(self, main, headers):
        super().__init__(main, headers)

        self.name = 'error'

class Wordless_Table_Data(Wordless_Table):
    def __init__(self, main,
                 headers, header_orientation = 'horizontal',
                 headers_int = [], headers_float = [],
                 headers_pct = [], headers_cumulative = [], cols_breakdown = [],
                 cols_stretch = [], sorting_enabled = False):
        super().__init__(main, headers, header_orientation, cols_stretch,
                         drag_drop_enabled = False)

        self.headers_int_old = headers_int
        self.headers_float_old = headers_float
        self.headers_pct_old = headers_pct
        self.headers_cumulative_old = headers_cumulative
        self.cols_breakdown_old = cols_breakdown

        self.sorting_enabled = sorting_enabled

        if sorting_enabled:
            self.setSortingEnabled(True)

            if header_orientation == 'horizontal':
                self.horizontalHeader().sortIndicatorChanged.connect(self.sorting_changed)
            else:
                self.verticalHeader().sortIndicatorChanged.connect(self.sorting_changed)

        self.itemChanged.connect(self.item_changed)
        self.itemSelectionChanged.connect(self.selection_changed)

        self.button_export_selected = QPushButton(self.tr('Export Selected...'), self)
        self.button_export_all = QPushButton(self.tr('Export All...'), self)
        self.button_clear = QPushButton(self.tr('Clear'), self)

        self.button_export_selected.clicked.connect(self.export_selected)
        self.button_export_all.clicked.connect(self.export_all)
        self.button_clear.clicked.connect(lambda: self.clear_table())

        self.clear_table()

    def item_changed(self):
        rows_visible = len([i for i in range(self.rowCount()) if not self.isRowHidden(i)])

        if [i for i in range(self.columnCount()) if self.item(0, i)] and rows_visible:
            self.button_export_all.setEnabled(True)
        else:
            self.button_export_all.setEnabled(False)

        super().item_changed()

        self.selection_changed()

    def selection_changed(self):
        if self.selectedIndexes() and [i for i in range(self.columnCount()) if self.item(0, i)]:
            self.button_export_selected.setEnabled(True)
        else:
            self.button_export_selected.setEnabled(False)

    def sorting_changed(self, logicalIndex, order):
        if [i for i in range(self.columnCount()) if self.item(0, i)]:
            self.update_ranks()

            if self.show_cumulative:
                self.toggle_cumulative()

    def insert_row(self, i, label,
                   is_int = False, is_float = False,
                   is_pct = False, is_cumulative = False, is_breakdown = False):
        headers_int = [self.verticalHeaderItem(row).text() for row in self.headers_int]
        headers_float = [self.verticalHeaderItem(row).text() for row in self.headers_float]
        headers_pct = [self.verticalHeaderItem(row).text() for row in self.headers_pct]
        headers_cumulative = [self.verticalHeaderItem(row).text() for row in self.headers_cumulative]
        cols_breakdown = [self.verticalHeaderItem(row).text() for row in self.cols_breakdown]

        super().insert_row(i, label)

        if is_int:
            headers_int += [label]
        if is_float:
            headers_float += [label]
        if is_pct:
            headers_pct += [label]
        if is_cumulative:
            headers_cumulative += [label]
        if is_breakdown:
            headers_cumulative += [label]

        self.headers_int = set(self.find_row(headers_int))
        self.headers_float = set(self.find_row(headers_float))
        self.headers_pct = set(self.find_row(headers_pct))
        self.headers_cumulative = set(self.find_row(headers_cumulative))
        self.cols_breakdown = set(self.find_row(cols_breakdown))

    def insert_col(self, i, label,
                   is_int = False, is_float = False,
                   is_pct = False, is_cumulative = False, is_breakdown = False):
        if self.header_orientation == 'horizontal':
            headers_int = [self.horizontalHeaderItem(col).text() for col in self.headers_int]
            headers_float = [self.horizontalHeaderItem(col).text() for col in self.headers_float]
            headers_pct = [self.horizontalHeaderItem(col).text() for col in self.headers_pct]
            headers_cumulative = [self.horizontalHeaderItem(col).text() for col in self.headers_cumulative]

        cols_breakdown = [self.horizontalHeaderItem(col).text() for col in self.cols_breakdown]

        super().insert_col(i, label)

        if is_int:
            headers_int += [label]
        if is_float:
            headers_float += [label]
        if is_pct:
            headers_pct += [label]
        if is_cumulative:
            headers_cumulative += [label]
        if is_breakdown:
            cols_breakdown += [label]

        if self.header_orientation == 'horizontal':
            self.headers_int = set(self.find_col(headers_int))
            self.headers_float = set(self.find_col(headers_float))
            self.headers_pct = set(self.find_col(headers_pct))
            self.headers_cumulative = set(self.find_col(headers_cumulative))
        
        self.cols_breakdown = set(self.find_col(cols_breakdown))

    def set_item_num(self, row, col, val, total = 0):
        if self.header_orientation == 'horizontal':
            header = col
        else:
            header = row

        # Integers
        if header in self.headers_int:
            val = int(val)

            item = Wordless_Table_Item(str(val))
        # Floats
        elif header in self.headers_float:
            val = float(val)
            precision = self.main.settings_custom['data']['precision_decimal']

            item = Wordless_Table_Item(f'{val:.{precision}f}')
        # Percentages
        elif header in self.headers_pct:
            # Handle zero division error
            if total:
                val = val / total
            else:
                val = 0

            precision = self.main.settings_custom['data']['precision_pct']

            item = Wordless_Table_Item(f'{val:.{precision}%}')

        item.val = val

        item.setFont(QFont('Consolas'))
        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)

        super().setItem(row, col, item)

    def update_ranks(self):
        data_prev = ''
        rank_prev = 1
        rank_next = 1

        sort_section = self.horizontalHeader().sortIndicatorSection()
        sort_order = self.horizontalHeader().sortIndicatorOrder()

        col_rank = self.find_col(self.tr('Rank'))

        self.sortItems(sort_section, sort_order)

        if sort_section != col_rank:
            self.blockSignals(True)
            self.setSortingEnabled(False)
            self.setUpdatesEnabled(False)

            for row in range(self.rowCount()):
                if not self.isRowHidden(row):
                    data_cur = self.item(row, sort_section).read_data()

                    if data_cur == data_prev:
                        self.item(row, col_rank).val = rank_prev
                        self.item(row, col_rank).setText(str(rank_prev))
                    else:
                        self.item(row, col_rank).val = rank_next
                        self.item(row, col_rank).setText(str(rank_next))

                        rank_prev = rank_next

                    rank_next += 1
                    data_prev = data_cur

            self.blockSignals(False)
            self.setSortingEnabled(True)
            self.setUpdatesEnabled(True)

    def toggle_pct(self):
        self.setUpdatesEnabled(False)

        if self.header_orientation == 'horizontal':
            if self.show_pct:
                for col in self.headers_pct:
                    self.showColumn(col)
            else:
                for col in self.headers_pct:
                    self.hideColumn(col)
        else:
            if self.show_pct:
                for row in self.headers_pct:
                    self.showRow(row)
            else:
                for row in self.headers_pct:
                    self.hideRow(row)

        self.setUpdatesEnabled(True)

    def toggle_cumulative(self):
        precision_decimal = self.main.settings_custom['data']['precision_decimal']
        precision_pct = self.main.settings_custom['data']['precision_pct']

        # Boost performance
        self.sortItems(self.horizontalHeader().sortIndicatorSection(), self.horizontalHeader().sortIndicatorOrder())

        self.blockSignals(True)
        self.setSortingEnabled(False)
        self.setUpdatesEnabled(False)

        if self.header_orientation == 'horizontal':
            if self.show_cumulative:
                for col in self.headers_cumulative:
                    val_cumulative = 0

                    # Integers
                    if col in self.headers_int:
                        for row in range(self.rowCount()):
                            if not self.isRowHidden(row):
                                item = self.item(row, col)

                                val_cumulative += item.val
                                item.setText(str(val_cumulative))
                    # Floats
                    elif col in self.headers_float:
                        for row in range(self.rowCount()):
                            if not self.isRowHidden(row):
                                item = self.item(row, col)

                                val_cumulative += item.val
                                item.setText(f'{val_cumulative:.{precision_decimal}}')
                    # Percentages
                    elif col in self.headers_pct:
                        for row in range(self.rowCount()):
                            if not self.isRowHidden(row):
                                item = self.item(row, col)

                                val_cumulative += item.val
                                item.setText(f'{val_cumulative:.{precision_pct}%}')
            else:
                for col in self.headers_cumulative:
                    # Integers
                    if col in self.headers_int:
                        for row in range(self.rowCount()):
                            if not self.isRowHidden(row):
                                item = self.item(row, col)

                                item.setText(str(item.val))
                    # Floats
                    elif col in self.headers_float:
                        for row in range(self.rowCount()):
                            if not self.isRowHidden(row):
                                item = self.item(row, col)

                                item.setText(f'{item.val:.{precision_decimal}}')
                    # Percentages
                    elif col in self.headers_pct:
                        for row in range(self.rowCount()):
                            if not self.isRowHidden(row):
                                item = self.item(row, col)

                                item.setText(f'{item.val:.{precision_pct}%}')
        else:
            if self.show_cumulative:
                for row in self.headers_cumulative:
                    val_cumulative = 0

                    # Integers
                    if row in self.headers_int:
                        for col in range(self.columnCount() - 1):
                            if not self.isColumnHidden(col):
                                item = self.item(row, col)

                                val_cumulative += item.val
                                item.setText(str(val_cumulative))
                    # Floats
                    elif row in self.headers_float:
                        for col in range(self.columnCount() - 1):
                            if not self.isColumnHidden(col):
                                item = self.item(row, col)

                                val_cumulative += item.val
                                item.setText(f'{val_cumulative:.{precision_decimal}}')
                    # Percentages
                    elif row in self.headers_pct:
                        for col in range(self.columnCount() - 1):
                            if not self.isColumnHidden(col):
                                item = self.item(row, col)

                                val_cumulative += item.val
                                item.setText(f'{val_cumulative:.{precision_pct}%}')
            else:
                for row in self.headers_cumulative:
                    # Integers
                    if row in self.headers_int:
                        for col in range(self.columnCount() - 1):
                            if not self.isColumnHidden(col):
                                item = self.item(row, col)

                                item.setText(str(item.val))
                    # Floats
                    elif row in self.headers_float:
                        for col in range(self.columnCount() - 1):
                            if not self.isColumnHidden(col):
                                item = self.item(row, col)

                                item.setText(f'{item.val:.{precision_decimal}}')
                    # Percentages
                    elif row in self.headers_pct:
                        for col in range(self.columnCount() - 1):
                            if not self.isColumnHidden(col):
                                item = self.item(row, col)

                                item.setText(f'{item.val:.{precision_pct}%}')

        self.blockSignals(False)
        self.setUpdatesEnabled(True)

        if self.sorting_enabled:
            self.setSortingEnabled(True)


    def toggle_breakdown(self):
        self.setUpdatesEnabled(False)

        if self.show_breakdown:
            for col in self.cols_breakdown:
                self.showColumn(col)
        else:
            for col in self.cols_breakdown:
                self.hideColumn(col)

        self.setUpdatesEnabled(True)

    def filter_table(self):
        rank_prev = 1
        rank_next = 1
        data_prev = ''

        self.hide()
        self.setUpdatesEnabled(False)
        
        for i, row_filter in enumerate(self.row_filters):
            if row_filter:
                self.showRow(i)
            else:
                self.hideRow(i)

        self.setUpdatesEnabled(True)
        self.show()

        self.toggle_cumulative()
        self.update_ranks()

        self.itemChanged.emit(self.item(0, 0))

    def clear_table(self, header_count = 1):
        self.clearContents()

        if self.header_orientation == 'horizontal':
            self.horizontalHeader().blockSignals(True)

            self.setColumnCount(len(self.headers))
            self.setRowCount(header_count)

            self.setHorizontalHeaderLabels(self.headers)

            self.horizontalHeader().blockSignals(False)

            self.horizontalHeader().sectionCountChanged.emit(0, header_count)
        else:
            self.verticalHeader().blockSignals(True)

            self.setRowCount(len(self.headers))
            self.setColumnCount(header_count)

            self.setVerticalHeaderLabels(self.headers)

            self.verticalHeader().blockSignals(False)

            self.verticalHeader().sectionCountChanged.emit(0, header_count)

        for i in range(self.rowCount()):
            self.showRow(i)

        for i in range(self.columnCount()):
            self.showColumn(i)

        self.headers_int = set(self.find_header(self.headers_int_old))
        self.headers_float = set(self.find_header(self.headers_float_old))
        self.headers_pct = set(self.find_header(self.headers_pct_old))
        self.headers_cumulative = set(self.find_header(self.headers_cumulative_old))
        self.cols_breakdown = set(self.find_col(self.cols_breakdown_old))

        self.item_changed()

class Wordless_Table_Data_Search(Wordless_Table_Data):
    def __init__(self, main, tab,
                 headers, header_orientation = 'horizontal',
                 headers_int = [], headers_float = [],
                 headers_pct = [], headers_cumulative = [], cols_breakdown = [],
                 cols_stretch = [], sorting_enabled = False):
        super().__init__(main,
                         headers, header_orientation,
                         headers_int, headers_float,
                         headers_pct, headers_cumulative, cols_breakdown,
                         cols_stretch, sorting_enabled)

        self.label_number_results = QLabel()
        self.button_results_search = wordless_button.Wordless_Button_Results_Search(self,
                                                                                    tab = tab,
                                                                                    table = self)

        self.itemChanged.connect(self.results_changed)

        self.results_changed()

    def results_changed(self):
        rows_visible = len([i for i in range(self.rowCount()) if not self.isRowHidden(i)])

        if [i for i in range(self.columnCount()) if self.item(0, i)] and rows_visible:
            self.label_number_results.setText(self.tr(f'Number of Results: {rows_visible}'))

            self.button_results_search.setEnabled(True)
        else:
            self.label_number_results.setText(self.tr('Number of Results: 0'))

            self.button_results_search.setEnabled(False)

class Wordless_Table_Data_Sort_Search(Wordless_Table_Data):
    def __init__(self, main, tab,
                 headers, header_orientation = 'horizontal',
                 headers_int = [], headers_float = [],
                 headers_pct = [], headers_cumulative = [], cols_breakdown = [],
                 cols_stretch = [], sorting_enabled = False):
        super().__init__(main,
                         headers, header_orientation,
                         headers_int, headers_float,
                         headers_pct, headers_cumulative, cols_breakdown,
                         cols_stretch, sorting_enabled)

        self.label_number_results = QLabel()
        self.button_results_sort = wordless_button.Wordless_Button_Results_Sort(self,
                                                                                table = self)
        self.button_results_search = wordless_button.Wordless_Button_Results_Search(self,
                                                                                    tab = tab,
                                                                                    table = self)

        self.itemChanged.connect(self.results_changed)

        self.results_changed()

    def results_changed(self):
        rows_visible = len([i for i in range(self.rowCount()) if not self.isRowHidden(i)])

        if [i for i in range(self.columnCount()) if self.item(0, i)] and rows_visible:
            self.label_number_results.setText(self.tr(f'Number of Results: {rows_visible}'))

            self.button_results_sort.setEnabled(True)
            self.button_results_search.setEnabled(True)
        else:
            self.label_number_results.setText(self.tr('Number of Results: 0'))

            self.button_results_sort.setEnabled(False)
            self.button_results_search.setEnabled(False)

class Wordless_Table_Data_Filter_Search(Wordless_Table_Data):
    def __init__(self, main, tab,
                 headers, header_orientation = 'horizontal',
                 headers_int = [], headers_float = [],
                 headers_pct = [], headers_cumulative = [], cols_breakdown = [],
                 cols_stretch = [], sorting_enabled = False):
        super().__init__(main,
                         headers, header_orientation,
                         headers_int, headers_float,
                         headers_pct, headers_cumulative, cols_breakdown,
                         cols_stretch, sorting_enabled)

        self.label_number_results = QLabel()
        self.button_results_filter = wordless_button.Wordless_Button_Results_Filter(self,
                                                                                    tab = tab,
                                                                                    table = self)
        self.button_results_search = wordless_button.Wordless_Button_Results_Search(self,
                                                                                    tab = tab,
                                                                                    table = self)

        self.itemChanged.connect(self.results_changed)

        self.results_changed()

    def results_changed(self):
        rows_visible = len([i for i in range(self.rowCount()) if not self.isRowHidden(i)])

        if [i for i in range(self.columnCount()) if self.item(0, i)] and rows_visible:
            self.label_number_results.setText(self.tr(f'Number of Results: {rows_visible}'))

            self.button_results_filter.setEnabled(True)
            self.button_results_search.setEnabled(True)
        else:
            self.label_number_results.setText(self.tr('Number of Results: 0'))

            self.button_results_filter.setEnabled(False)
            self.button_results_search.setEnabled(False)

class Wordless_Worker_Results_Sort_Concordancer(wordless_threading.Wordless_Worker):
    worker_done = pyqtSignal(list)

    def run(self):
        results = []

        len_left = max([
            int(self.dialog.cellWidget(0, 0).itemText(i)[1:])
            for i in range(self.dialog.cellWidget(0, 0).count())
            if 'L' in self.dialog.cellWidget(0, 0).itemText(i)]
        )
        len_right = max([
            int(self.dialog.cellWidget(0, 0).itemText(i)[1:])
            for i in range(self.dialog.cellWidget(0, 0).count())
            if 'R' in self.dialog.cellWidget(0, 0).itemText(i)
        ])

        for i in range(self.dialog.table.rowCount()):
            left_old = self.dialog.table.cellWidget(i, 0)
            node_old = self.dialog.table.cellWidget(i, 1)
            right_old = self.dialog.table.cellWidget(i, 2)

            if len(left_old.text_raw) < len_left:
                left_old.text_raw = [''] * (len_left - len(left_old.text_raw)) + left_old.text_raw
            if len(right_old.text_raw) < len_right:
                right_old.text_raw.extend([''] * (len_right - len(right_old.text_raw)))

            no_token = self.dialog.table.item(i, 3).val
            no_token_pct = self.dialog.table.item(i, 4).val
            no_clause = self.dialog.table.item(i, 5).val
            no_clause_pct = self.dialog.table.item(i, 6).val
            no_sentence = self.dialog.table.item(i, 7).val
            no_sentence_pct = self.dialog.table.item(i, 8).val
            no_para = self.dialog.table.item(i, 9).val
            no_para_pct = self.dialog.table.item(i, 10).val
            file = self.dialog.table.item(i, 11).text()

            results.append([
                left_old, node_old, right_old,
                no_token, no_token_pct,
                no_clause, no_clause_pct,
                no_sentence, no_sentence_pct,
                no_para, no_para_pct,
                file
            ])

        self.progress_updated.emit(self.tr('Updating table ...'))

        time.sleep(0.1)

        self.worker_done.emit(results)

class Wordless_Table_Results_Sort_Conordancer(Wordless_Table):
    def __init__(self, parent, table):
        super().__init__(parent,
                         headers = [
                             parent.tr('Columns'),
                             parent.tr('Order')
                         ],
                         cols_stretch = [
                             parent.tr('Order')
                         ])

        self.table = table
        self.cols_sorting = []

        self.button_add = QPushButton(self.tr('Add'), self)
        self.button_insert = QPushButton(self.tr('Insert'), self)
        self.button_remove = QPushButton(self.tr('Remove'), self)
    
        self.button_add.clicked.connect(self.add_row)
        self.button_insert.clicked.connect(self.insert_row)
        self.button_remove.clicked.connect(self.remove_row)

        self.itemChanged.connect(self.item_changed)
        self.itemSelectionChanged.connect(self.selection_changed)

        self.table.itemChanged.connect(self.table_item_changed)

        self.setFixedHeight(160)

        self.clear_table()

    def item_changed(self):
        if self.rowCount() < len(self.cols_sorting):
            self.button_add.setEnabled(True)
        else:
            self.button_add.setEnabled(False)

        for i in range(self.rowCount()):
            self.cellWidget(i, 0).text_old = self.cellWidget(i, 0).currentText()

        self.selection_changed()

    def selection_changed(self):
        if self.selectedIndexes() and self.rowCount() < len(self.cols_sorting):
            self.button_insert.setEnabled(True)
        else:
            self.button_insert.setEnabled(False)

        if self.selectedIndexes() and len(self.get_selected_rows()) < self.rowCount():
            self.button_remove.setEnabled(True)
        else:
            self.button_remove.setEnabled(False)

    def table_item_changed(self):
        sorting_rules = copy.deepcopy(self.main.settings_custom['concordancer']['sort_results']['sorting_rules'])

        self.setRowCount(0)

        # Columns to sort
        self.cols_sorting = [
            self.tr('Node'),
            self.tr('Token No.'),
            self.tr('File')
        ]

        if [i for i in range(self.table.columnCount()) if self.table.item(0, i)]:
            if self.table.settings['concordancer']['generation_settings']['width_unit'] == self.tr('Token'):
                width_left = self.table.settings['concordancer']['generation_settings']['width_left_token']
                width_right = self.table.settings['concordancer']['generation_settings']['width_right_token']
            else:
                col_left = self.table.find_col(self.tr('Left'))
                col_right = self.table.find_col(self.tr('Right'))

                width_left = max([len(self.table.cellWidget(row, col_left).text_raw)
                                  for row in range(self.table.rowCount())])
                width_right = max([len(self.table.cellWidget(row, col_right).text_raw)
                                   for row in range(self.table.rowCount())])

            self.cols_sorting.extend([f'R{i + 1}' for i in range(width_right)])
            self.cols_sorting.extend([f'L{i + 1}' for i in range(width_left)])

        # Check sorting settings
        for sorting_col, sorting_order in sorting_rules:
            if sorting_col in self.cols_sorting:
                self.add_row()

                self.cellWidget(self.rowCount() - 1, 0).setCurrentText(sorting_col)
                self.cellWidget(self.rowCount() - 1, 1).setCurrentText(sorting_order)

        self.itemChanged.emit(self.item(0, 0))

    def sorting_col_changed(self, combo_box_sorting_col):
        for i in range(self.rowCount()):
            combo_box_cur = self.cellWidget(i, 0)

            if combo_box_sorting_col != combo_box_cur and combo_box_sorting_col.currentText() == combo_box_cur.currentText():
                QMessageBox.warning(self.main,
                                    self.tr('Column Sorted More Than Once'),
                                    self.tr(f'''
                                        {self.main.settings_global['styles']['style_dialog']}
                                        <body>
                                            <div>Please refrain from sorting the same column more than once!</div>
                                        </body>
                                    '''),
                                    QMessageBox.Ok)

                combo_box_sorting_col.setCurrentText(combo_box_sorting_col.text_old)
                combo_box_sorting_col.showPopup()

                return

        combo_box_sorting_col.text_old = combo_box_sorting_col.currentText()

    def _new_row(self):
        combo_box_sorting_col = wordless_box.Wordless_Combo_Box(self)
        combo_box_sorting_order = wordless_box.Wordless_Combo_Box(self)

        combo_box_sorting_col.addItems(self.cols_sorting)
        combo_box_sorting_order.addItems([
            self.tr('Ascending'),
            self.tr('Descending')
        ])

        if combo_box_sorting_col.findText('L1') > -1:
            width_left = max([int(combo_box_sorting_col.itemText(i)[1:])
                              for i in range(combo_box_sorting_col.count())
                              if 'L' in combo_box_sorting_col.itemText(i)])
        else:
            width_left = 0

        if combo_box_sorting_col.findText('R1') > -1:
            width_right = max([int(combo_box_sorting_col.itemText(i)[1:])
                               for i in range(combo_box_sorting_col.count())
                               if 'R' in combo_box_sorting_col.itemText(i)])
        else:
            width_right = 0

        cols_left = [int(self.cellWidget(i, 0).currentText()[1:])
                     for i in range(self.rowCount())
                     if 'L' in self.cellWidget(i, 0).currentText()]
        cols_right = [int(self.cellWidget(i, 0).currentText()[1:])
                      for i in range(self.rowCount())
                      if 'R' in self.cellWidget(i, 0).currentText()]

        if cols_left and max(cols_left) < width_left:
            combo_box_sorting_col.setCurrentText(f'L{cols_left[-1] + 1}')
        elif cols_right and max(cols_right) < width_right:
            combo_box_sorting_col.setCurrentText(f'R{cols_right[-1] + 1}')
        elif cols_right and max(cols_right) and not cols_left:
            combo_box_sorting_col.setCurrentText(f'L1')
        else:
            for i in range(combo_box_sorting_col.count()):
                text = combo_box_sorting_col.itemText(i)

                if text not in [self.cellWidget(j, 0).currentText() for j in range(self.rowCount())]:
                    combo_box_sorting_col.setCurrentText(text)

                    break

        combo_box_sorting_col.currentTextChanged.connect(lambda: self.sorting_col_changed(combo_box_sorting_col))
        combo_box_sorting_col.currentTextChanged.connect(lambda: self.itemChanged.emit(self.item(0, 0)))
        combo_box_sorting_order.currentTextChanged.connect(lambda: self.itemChanged.emit(self.item(0, 0)))

        return (combo_box_sorting_col, combo_box_sorting_order)

    def add_row(self):
        combo_box_sorting_col, combo_box_sorting_order = self._new_row()
        
        self.setRowCount(self.rowCount() + 1)
        self.setCellWidget(self.rowCount() - 1, 0, combo_box_sorting_col)
        self.setCellWidget(self.rowCount() - 1, 1, combo_box_sorting_order)

        self.selectRow(self.rowCount() - 1)

        self.itemChanged.emit(self.item(0, 0))

    def insert_row(self):
        row = self.get_selected_rows()[0]

        combo_box_sorting_col, combo_box_sorting_order = self._new_row()

        self.insertRow(row)

        self.setCellWidget(row, 0, combo_box_sorting_col)
        self.setCellWidget(row, 1, combo_box_sorting_order)

        self.selectRow(row)

        self.itemChanged.emit(self.item(0, 0))

    def remove_row(self):
        for i in reversed(self.get_selected_rows()):
            self.removeRow(i)

        self.itemChanged.emit(self.item(0, 0))

    @wordless_misc.log_timing
    def sort_results(self):
        def update_gui(results):
            # Create new labels
            for i, (left_old, node_old, right_old,
                    _, _, _, _, _, _, _, _, _) in enumerate(results):
                left_new = wordless_label.Wordless_Label_Html('', self.table)
                node_new = wordless_label.Wordless_Label_Html(node_old.text(), self.table)
                right_new = wordless_label.Wordless_Label_Html('', self.table)

                left_new.text_raw = left_old.text_raw.copy()
                node_new.text_raw = node_old.text_raw.copy()
                right_new.text_raw = right_old.text_raw.copy()

                left_new.text_search = left_old.text_search.copy()
                node_new.text_search = node_old.text_search.copy()
                right_new.text_search = right_old.text_search.copy()

                results[i][0] = left_new
                results[i][1] = node_new
                results[i][2] = right_new

            self.table.blockSignals(True)
            self.table.setUpdatesEnabled(False)

            for i, (left, node, right,
                    no_token, no_token_pct,
                    no_clause, no_clause_pct,
                    no_sentence, no_sentence_pct,
                    no_para, no_para_pct,
                    file) in enumerate(sorted(results, key = key_concordancer)):
                for file_open in self.table.settings['files']['files_open']:
                    if file_open['selected'] and file_open['name'] == file:
                        lang = file_open['lang']

                # Remove empty tokens
                text_left = [token for token in left.text_raw if token]
                text_right = [token for token in right.text_raw if token]

                highlight_colors = self.main.settings_custom['concordancer']['sort_results']['highlight_colors']

                i_highlight_color_left = 1
                i_highlight_color_right = 1

                for j, key in enumerate([key for key in sorting_keys if type(key) != int]):
                    if key[0] == 0 and -key[1] <= len(text_left):
                        hightlight_color = highlight_colors[i_highlight_color_left % len(highlight_colors)]

                        text_left[key[1]] = f'''
                            <span style="color: {hightlight_color}; font-weight: bold;">
                                {text_left[key[1]]}
                            </span>
                        '''

                        i_highlight_color_left += 1
                    elif key[0] == 2 and key[1] < len(text_right):
                        hightlight_color = highlight_colors[i_highlight_color_right % len(highlight_colors)]

                        text_right[key[1]] = f'''
                            <span style="color: {hightlight_color}; font-weight: bold;">
                                {text_right[key[1]]}
                            </span>
                        '''

                        i_highlight_color_right += 1

                text_left = wordless_text_processing.wordless_word_detokenize(self.main, text_left, lang)
                text_right = wordless_text_processing.wordless_word_detokenize(self.main, text_right, lang)

                self.table.cellWidget(i, 0).setText(text_left)
                self.table.cellWidget(i, 1).setText(node.text())
                self.table.cellWidget(i, 2).setText(text_right)

                self.table.cellWidget(i, 0).text_raw = [token for token in left.text_raw if token]
                self.table.cellWidget(i, 1).text_raw = node.text_raw
                self.table.cellWidget(i, 2).text_raw = [token for token in right.text_raw if token]

                self.table.cellWidget(i, 0).text_search = left.text_search
                self.table.cellWidget(i, 1).text_search = node.text_search
                self.table.cellWidget(i, 2).text_search = right.text_search

                self.table.set_item_num(i, 3, no_token)
                self.table.set_item_num(i, 4, no_token_pct)
                self.table.set_item_num(i, 5, no_clause)
                self.table.set_item_num(i, 6, no_clause_pct)
                self.table.set_item_num(i, 7, no_sentence)
                self.table.set_item_num(i, 8, no_sentence_pct)
                self.table.set_item_num(i, 9, no_para)
                self.table.set_item_num(i, 10, no_para_pct)
                self.table.item(i, 11).setText(file)

            self.table.setUpdatesEnabled(True)
            self.table.blockSignals(False)

        def key_concordancer(item):
            keys = []

            for key in sorting_keys:
                # Node
                if key == 1:
                    keys.append(item[key].text_raw)
                # Left & Right
                elif type(key) == list:
                    keys.append(item[key[0]].text_raw[key[1]])
                else:
                    keys.append(item[key])

            return keys

        sorting_keys = []

        settings = self.table.settings['concordancer']

        if [i for i in range(self.table.columnCount()) if self.table.item(0, i)]:
            for sorting_col, sorting_order in settings['sort_results']['sorting_rules']:
                if sorting_col == self.tr('File'):
                    sorting_keys.append(11)
                elif sorting_col == self.tr('Token No.'):
                    sorting_keys.append(3)
                elif sorting_col == self.tr('Node'):
                    sorting_keys.append(1)
                elif 'R' in sorting_col:
                    sorting_keys.append([2, int(sorting_col[1:]) - 1])
                elif 'L' in sorting_col:
                    sorting_keys.append([0, -int(sorting_col[1:])])

            dialog_progress = wordless_dialog_misc.Wordless_Dialog_Progress_Results_Sort(self.main)

            worker_results_sort_concordancer = Wordless_Worker_Results_Sort_Concordancer(
                self.main,
                dialog_progress = dialog_progress,
                update_gui = update_gui,
                dialog = self
            )

            thread_results_sort_concordancer = wordless_threading.Wordless_Thread(worker_results_sort_concordancer)
            thread_results_sort_concordancer.start_worker()

        wordless_msg.wordless_msg_results_sort(self.main)

class Wordless_Table_Tags(Wordless_Table):
    def __init__(self, main):
        super().__init__(main,
                         headers = [
                             main.tr('Opening Tag'),
                             main.tr('Closing Tag'),
                             main.tr('Preview')
                         ],
                         header_orientation = 'horizontal',
                         cols_stretch = [
                             main.tr('Preview')
                         ],
                         drag_drop_enabled = True)

        self.verticalHeader().setHidden(True)
        self.setFixedHeight(125)

        self.setEditTriggers(QAbstractItemView.DoubleClicked | QAbstractItemView.SelectedClicked)

        self.itemChanged.connect(self.item_changed)

        self.button_add = QPushButton(self.tr('Add'), self)
        self.button_remove = QPushButton(self.tr('Remove'), self)
        self.button_reset = QPushButton(self.tr('Reset'), self)

        self.button_add.clicked.connect(self.add_item)
        self.button_remove.clicked.connect(self.remove_item)
        self.button_reset.clicked.connect(self.reset_table)

        self.reset_table()

    def item_changed(self, item = None):
        self.blockSignals(True)

        if item:
            # Opening Tag
            if item.column() == 0:
                if re.search(r'^\s*$', item.text()):
                    QMessageBox.warning(self.main,
                                        self.tr('Empty Opening Tag'),
                                        self.tr(f'''
                                            {self.main.settings_global['styles']['style_dialog']}
                                            <body>
                                                <p>The opening tag should not be left empty!</p>
                                            </body>
                                        '''))

                    item.setText(item.text_old)

                    self.closePersistentEditor(item)
                    self.editItem(item)

                    return
            # Closing Tag
            elif item.column() == 1:
                if re.search(r'^\s*$', item.text()):
                    item.setText('')

            for row in range(self.rowCount()):
                if row != item.row():
                    if (self.item(row, 0).text() == self.item(item.row(), 0).text() and
                        self.item(row, 1).text() == self.item(item.row(), 1).text()):
                        wordless_msg_box.wordless_msg_box_duplicate_tags(self.main)

                        item.setText(item.text_old)

                        self.closePersistentEditor(item)
                        self.editItem(item)

                        return
        
        for row in range(self.rowCount()):
            for col in range(self.columnCount()):
                if col < 2:
                    self.item(row, col).text_old = self.item(row, col).text()

            self.item(row, 2).setText(self.tr(f'token{self.item(row, 0).text()}TAG{self.item(row, 1).text()}'))

        self.blockSignals(False)

    def _new_item(self, text = None):
        i = 1

        if text == None:
            while True:
                if self.findItems(self.tr(f'New Tag ({i})'), Qt.MatchExactly):
                    i += 1
                else:
                    new_item = QTableWidgetItem(self.tr('New Tag ({})').format(i))

                    break
        else:
            new_item = QTableWidgetItem(text)

        new_item.text_old = new_item.text()
        
        return new_item

    def add_item(self, texts = []):
        self.blockSignals(True)

        self.setRowCount(self.rowCount() + 1)

        if texts:
            for i in range(self.columnCount()):
                if i < 2:
                    self.setItem(self.rowCount() - 1, i, self._new_item(text = texts[i]))
                else:
                    self.setItem(self.rowCount() - 1, i, QTableWidgetItem(''))
        else:
            for i in range(self.columnCount()):
                if i < 1:
                    self.setItem(self.rowCount() - 1, i, self._new_item())
                else:
                    self.setItem(self.rowCount() - 1, i, QTableWidgetItem(''))

                self.item(self.rowCount() - 1, i).setSelected(True)

            self.editItem(self.item(self.rowCount() - 1, 0))

        self.item(self.rowCount() - 1, self.columnCount() - 1).setFlags(Qt.ItemIsSelectable |
                                                                        Qt.ItemIsDragEnabled |
                                                                        Qt.ItemIsEnabled)

        self.blockSignals(False)

        self.item_changed()

    def remove_item(self):
        if len(self.get_selected_rows()) == self.rowCount():
            QMessageBox.warning(self.main,
                                self.tr('Empty Tag'),
                                self.tr(f'You should specify at least 1 (pair of) tag in the table!'))
        else:
            self.blockSignals(True)

            for i in reversed(self.get_selected_rows()):
                self.removeRow(i)

            self.blockSignals(False)

    def reset_table(self):
        self.blockSignals(True)

        self.setRowCount(0)

        self.blockSignals(False)

    def get_tags(self):
        tags = []

        for row in range(self.rowCount()):
            tags.append([self.item(row, 0).text(),
                         self.item(row, 1).text()])

        return tags
